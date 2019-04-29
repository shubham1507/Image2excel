from PIL import Image
import re
from pytesseract import image_to_string
import os
import xlwt
from xlwt import Workbook 
import openpyxl 



# wb = Workbook() 
wb = openpyxl.Workbook()


# sheet1 = wb.add_sheet('Sheet 1') 

sheet = wb.active


path = '//home/snj/Downloads/temp/'
temp_list = []
for filename in os.listdir(path):

    temp = image_to_string(Image.open('/home/snj/Downloads/temp/'+filename), lang='eng')
    temp1 = str([int(s) for s in temp.split() if s.isdigit()])
    temp1 = (re.findall('996\d+',temp1))
    temp1 = [s for s in temp1 if len(s)==10]

    for i in temp1:
        temp_list.append(i)


print((temp_list))

le_ = len(temp_list)

p = 0

for i in temp_list:
    k = 0
    for j in i:
        c1 = sheet.cell(row=p+1,column=k+1)
        
        c1.value = str(i)
        
        
    p+=1







wb.save("demo1.xlsx")




 
